#!/usr/bin/env python3
"""
Génère un flux RSS (feed.xml) à partir de la diet du jour stockée dans
le fichier Google Drive "Coaching DODA".

Fonctionnement :
1. Télécharge le fichier .xlsx depuis Google Drive (le fichier doit être
   partagé en "Toute personne disposant du lien peut consulter").
2. Cherche la date du jour dans la feuille "2026 DIET (2)".
3. Formate le plan alimentaire du jour en HTML.
4. Ajoute une entrée <item> au flux RSS (feed.xml), en conservant les
   entrées des jours précédents (jusqu'à MAX_ITEMS).

Variables d'environnement attendues :
- DRIVE_FILE_ID   : ID du fichier Google Drive (obligatoire)
- SHEET_NAME      : nom de la feuille à utiliser (par défaut "2026 DIET (2)")
- FEED_TITLE      : titre du flux RSS (par défaut "Diet du jour - DODA")
- FEED_LINK       : URL publique du flux (ex: https://<user>.github.io/<repo>/feed.xml)
- TIMEZONE        : fuseau horaire pour déterminer "aujourd'hui" (par défaut "Europe/Brussels")
"""

import os
import sys
import datetime
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape

import gdown
import openpyxl

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    from backports.zoneinfo import ZoneInfo  # type: ignore

DRIVE_FILE_ID = os.environ.get("DRIVE_FILE_ID", "1nGm7eU-hB_qvJKOUhZK4reLliot20HIyYDxMwIMPU8I")
SHEET_NAME = os.environ.get("SHEET_NAME", "2026 DIET (2)")
FEED_TITLE = os.environ.get("FEED_TITLE", "Diet du jour - DODA")
FEED_LINK = os.environ.get("FEED_LINK", "https://example.github.io/diet-rss-feed/feed.xml")
FEED_DESCRIPTION = "Plan alimentaire quotidien généré automatiquement"
TIMEZONE = os.environ.get("TIMEZONE", "Europe/Brussels")
MAX_ITEMS = 14

XLSX_PATH = "diet.xlsx"
FEED_PATH = "feed.xml"


def download_xlsx(file_id: str, dest: str) -> None:
    """Télécharge un fichier Drive public (lien 'anyone with the link') via gdown.

    gdown gère nativement la page de confirmation Google pour les gros fichiers
    (l'ancienne méthode manuelle via `requests` plantait avec une erreur 500)."""
    output = gdown.download(id=file_id, output=dest, quiet=False)
    if output is None:
        raise RuntimeError(
            f"Échec du téléchargement du fichier Drive {file_id}. "
            "Vérifie que le fichier est bien partagé en 'Tous les utilisateurs "
            "disposant du lien' (rôle Lecteur)."
        )


def find_today_block(ws, target_date: datetime.date):
    """Trouve la cellule contenant la date cible, retourne (col, row)."""
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, (datetime.datetime, datetime.date)):
                cell_date = val.date() if isinstance(val, datetime.datetime) else val
                if cell_date == target_date:
                    return cell.column, cell.row
    return None, None


def extract_day(ws, col: int, row: int):
    """Extrait ~38 lignes x 7 colonnes à partir du bloc trouvé."""
    data = []
    for r in range(row, row + 39):
        line = []
        for c in range(col, col + 7):
            line.append(ws.cell(row=r, column=c).value)
        data.append(line)
    return data


def format_meals_html(rows) -> str:
    """Transforme les lignes brutes en HTML lisible, groupé par repas."""
    meal_headers = {"BREAKFAST", "LUNCH", "PRE WO", "POST WO", "DINER", "DÎNER", "SNACK"}

    # On regroupe d'abord les lignes par repas
    meals = []  # list of dicts: {name, header_kcal, header_g, header_l, header_p, items: [...]}
    current = None

    for line in rows[1:]:  # la 1ère ligne est le total du jour
        aliment, dose, grammes, kcal, g, l, p = (line + [None] * 7)[:7]

        if isinstance(aliment, (datetime.datetime, datetime.date)):
            break  # fin du bloc du jour, on entre dans le jour suivant

        if aliment is None and dose is None:
            continue

        label = str(aliment).strip().upper() if aliment else ""

        if label in meal_headers:
            current = {
                "name": str(aliment).strip(),
                "header_kcal": kcal if isinstance(kcal, (int, float)) else 0,
                "header_g": g if isinstance(g, (int, float)) else 0,
                "header_l": l if isinstance(l, (int, float)) else 0,
                "header_p": p if isinstance(p, (int, float)) else 0,
                "items": [],
            }
            meals.append(current)
            continue

        if current is None:
            continue  # ligne orpheline avant le premier repas identifié

        if aliment is None and dose is not None:
            current["items"].append({"text": str(dose), "kcal": None})
            continue

        if aliment is None:
            continue

        current["items"].append({
            "text": str(aliment),
            "dose": dose,
            "kcal": kcal if isinstance(kcal, (int, float)) else None,
            "g": g if isinstance(g, (int, float)) else 0,
            "l": l if isinstance(l, (int, float)) else 0,
            "p": p if isinstance(p, (int, float)) else 0,
        })

    html_parts = []
    total_kcal = total_g = total_l = total_p = 0.0

    for meal in meals:
        items_kcal_sum = sum(i["kcal"] for i in meal["items"] if i.get("kcal"))
        items_g_sum = sum(i.get("g", 0) for i in meal["items"])
        items_l_sum = sum(i.get("l", 0) for i in meal["items"])
        items_p_sum = sum(i.get("p", 0) for i in meal["items"])

        # Si les items ne sont pas chiffrés (repas "guideline"), on retombe sur le total d'en-tête
        if items_kcal_sum > 0:
            meal_kcal = items_kcal_sum
            total_g += items_g_sum
            total_l += items_l_sum
            total_p += items_p_sum
        else:
            meal_kcal = meal["header_kcal"]
            total_g += meal["header_g"]
            total_l += meal["header_l"]
            total_p += meal["header_p"]

        total_kcal += meal_kcal

        title = escape(meal["name"])
        if meal_kcal:
            html_parts.append(f"<h3>{title} <em>(~{meal_kcal:.0f} kcal)</em></h3><ul>")
        else:
            html_parts.append(f"<h3>{title}</h3><ul>")

        for item in meal["items"]:
            if "dose" not in item:
                html_parts.append(f"<li>{escape(item['text'])}</li>")
                continue
            dose_str = f" — {item['dose']}" if item["dose"] else ""
            kcal_str = f" → {item['kcal']:.0f} kcal" if item["kcal"] else ""
            html_parts.append(f"<li>{escape(str(item['text']))}{escape(dose_str)}{kcal_str}</li>")

        html_parts.append("</ul>")

    summary = (
        f"<p><strong>Total estimé : ~{total_kcal:.0f} kcal | "
        f"G: {total_g:.0f}g | L: {total_l:.0f}g | P: {total_p:.0f}g</strong></p>"
    )
    return summary + "".join(html_parts)


def load_existing_items(feed_path: str):
    """Récupère les <item> existants du feed.xml précédent, si présent."""
    if not os.path.exists(feed_path):
        return []
    try:
        tree = ET.parse(feed_path)
        channel = tree.getroot().find("channel")
        if channel is None:
            return []
        return channel.findall("item")
    except ET.ParseError:
        return []


def build_feed(new_title, new_link, new_guid, new_pubdate, new_description_html):
    existing_items = load_existing_items(FEED_PATH)

    # Évite les doublons si le script tourne plusieurs fois le même jour
    existing_items = [it for it in existing_items if (it.findtext("guid") or "") != new_guid]

    rss = ET.Element("rss", version="2.0")
    channel = ET.SubElement(rss, "channel")
    ET.SubElement(channel, "title").text = FEED_TITLE
    ET.SubElement(channel, "link").text = FEED_LINK
    ET.SubElement(channel, "description").text = FEED_DESCRIPTION
    ET.SubElement(channel, "lastBuildDate").text = datetime.datetime.now(
        datetime.timezone.utc
    ).strftime("%a, %d %b %Y %H:%M:%S %z")

    new_item = ET.Element("item")
    ET.SubElement(new_item, "title").text = new_title
    ET.SubElement(new_item, "link").text = new_link
    ET.SubElement(new_item, "guid", isPermaLink="false").text = new_guid
    ET.SubElement(new_item, "pubDate").text = new_pubdate
    desc = ET.SubElement(new_item, "description")
    desc.text = new_description_html

    channel.append(new_item)
    for it in existing_items[: MAX_ITEMS - 1]:
        channel.append(it)

    tree = ET.ElementTree(rss)
    ET.indent(tree, space="  ")
    tree.write(FEED_PATH, encoding="utf-8", xml_declaration=True)


def main():
    tz = ZoneInfo(TIMEZONE)
    today = datetime.datetime.now(tz).date()

    print(f"Téléchargement du fichier Drive {DRIVE_FILE_ID}...")
    download_xlsx(DRIVE_FILE_ID, XLSX_PATH)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERREUR: feuille '{SHEET_NAME}' introuvable. Feuilles disponibles: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    col, row = find_today_block(ws, today)
    if col is None:
        print(f"Aucune date correspondant à {today} trouvée dans '{SHEET_NAME}'.")
        description_html = (
            f"<p>Aucune diet trouvée pour le {today.strftime('%d/%m/%Y')}. "
            "Le fichier n'a peut-être pas encore été mis à jour.</p>"
        )
    else:
        day_rows = extract_day(ws, col, row)
        description_html = format_meals_html(day_rows)

    title = f"Diet du {today.strftime('%d/%m/%Y')}"
    guid = f"diet-{today.isoformat()}"
    pubdate = datetime.datetime.now(datetime.timezone.utc).strftime("%a, %d %b %Y %H:%M:%S %z")

    build_feed(title, FEED_LINK, guid, pubdate, description_html)
    print(f"Flux RSS mis à jour : {FEED_PATH}")


if __name__ == "__main__":
    main()
